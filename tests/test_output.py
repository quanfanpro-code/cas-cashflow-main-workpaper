from pathlib import Path

import openpyxl

from cashflow_main.completeness_check import ValidationReport
from cashflow_main.contracts import EnterpriseType, InputManifest
from cashflow_main.item_calculators import CalculationResult, CashflowItemResult
from cashflow_main.output import write_cashflow_statement
from cashflow_main.pipeline import RunConfig


def test_output_has_statement_and_visible_verification_workpaper(tmp_path):
    items = (
        CashflowItemResult("CFO-01", "销售商品、提供劳务收到的现金", "operating", 10, 12345, ()),
        CashflowItemResult("CFO-NET", "经营活动产生的现金流量净额", "operating", 20, 12345, ()),
    )
    dummy = tmp_path / "x"
    manifest = InputManifest(dummy, dummy, dummy, dummy, dummy, "元", "人民币", 100)
    config = RunConfig(manifest, EnterpriseType.GENERAL, "测试公司", "2025年度")
    validation = ValidationReport((), (), (), 0)
    path = write_cashflow_statement(CalculationResult(items, {}), {"CFO-01": 99.5}, config, validation, "最终", tmp_path / "现金流量表.xlsx")
    wb = openpyxl.load_workbook(path, data_only=False)
    assert wb.sheetnames == ["现金流量表", "计算验证底稿"]
    ws = wb["现金流量表"]
    assert ws.sheet_state == "visible"
    assert ws["C6"].value == 123.45
    assert ws["D6"].value == 99.5
    assert ws["B6"].value is None
    assert not any(isinstance(cell.value, str) and cell.value.startswith("=") for row in ws.iter_rows() for cell in row)
    verification = wb["计算验证底稿"]
    assert verification.sheet_state == "visible"
    assert any(isinstance(cell.value, str) and cell.value.startswith("=") for row in verification.iter_rows() for cell in row)
    assert not wb._external_links


def test_unmatched_prior_item_is_visible_in_workpaper(tmp_path):
    item = CashflowItemResult("CFO-01", "销售商品、提供劳务收到的现金", "operating", 10, 100, ())
    dummy = tmp_path / "x"
    config = RunConfig(InputManifest(dummy, dummy, dummy, dummy, dummy, "元", "人民币", 100), EnterpriseType.GENERAL)
    path = write_cashflow_statement(
        CalculationResult((item,), {}), {}, config, ValidationReport((), (), (), 0), "最终", tmp_path / "未匹配.xlsx"
    )
    wb = openpyxl.load_workbook(path, data_only=False)
    values = [cell.value for row in wb["计算验证底稿"].iter_rows() for cell in row]
    assert "上期项目匹配" in values
    assert "销售商品、提供劳务收到的现金" in values
    assert "未匹配，上期数暂列0" in values


def test_provisional_output_is_visibly_marked(tmp_path):
    dummy = tmp_path / "x"
    config = RunConfig(InputManifest(dummy, dummy, dummy, dummy, dummy, "元", "人民币", 100), EnterpriseType.BANK, "银行A", "2025年度")
    path = write_cashflow_statement(CalculationResult((), {}), {}, config, ValidationReport((), (), (), 0), "暂编", tmp_path / "暂编.xlsx")
    ws = openpyxl.load_workbook(path).active
    assert "暂编" in ws["A2"].value
