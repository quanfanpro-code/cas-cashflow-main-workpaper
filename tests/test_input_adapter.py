from pathlib import Path

import pytest
from openpyxl import Workbook

from cashflow_main.contracts import InputManifest
from cashflow_main.input_adapter import (
    InputFormulaCacheError,
    InputValidationError,
    normalize_inputs,
    relevant_sheet_names,
)


def write_table(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def build_manifest(tmp_path: Path) -> InputManifest:
    balance_sheet = tmp_path / "资产负债表.xlsx"
    income_statement = tmp_path / "利润表.xlsx"
    trial_balance = tmp_path / "科目余额表.xlsx"
    journal = tmp_path / "一借一贷明细.xlsx"
    prior_cashflow = tmp_path / "上期现金流量表.xlsx"
    write_table(
        balance_sheet,
        ["项目", "期末余额", "期初余额"],
        [["货币资金", 15000, 10000]],
    )
    write_table(
        income_statement,
        ["项目", "本期金额"],
        [["营业收入", 1000]],
    )
    write_table(
        trial_balance,
        ["科目编码", "科目名称", "期初余额", "借方发生额", "贷方发生额", "期末余额"],
        [["1002", "银行存款", 10000, 1000, 0, 11000]],
    )
    write_table(
        journal,
        ["凭证号", "借方科目", "贷方科目", "配对金额"],
        [["记-1", "银行存款", "应收账款", 1000]],
    )
    write_table(
        prior_cashflow,
        ["项目", "本期金额"],
        [["销售商品、提供劳务收到的现金", 900]],
    )
    return InputManifest(
        audited_balance_sheet_path=balance_sheet,
        audited_income_statement_path=income_statement,
        trial_balance_path=trial_balance,
        journal_pairs_path=journal,
        prior_cashflow_path=prior_cashflow,
        display_unit="元",
        currency="CNY",
        performance_materiality_minor=100_000,
    )


def test_normalize_requires_only_three_new_journal_fields(tmp_path):
    bundle = normalize_inputs(build_manifest(tmp_path))
    pair = bundle.journal_pairs[0]
    assert pair.debit_account_name == "银行存款"
    assert pair.credit_account_name == "应收账款"
    assert pair.amount_minor == 100_000
    assert pair.original_fields["凭证号"] == "记-1"


def test_missing_paired_amount_is_a_specific_input_issue(tmp_path):
    manifest = build_manifest(tmp_path)
    write_table(
        manifest.journal_pairs_path,
        ["凭证号", "借方科目", "贷方科目"],
        [["记-1", "银行存款", "应收账款"]],
    )
    with pytest.raises(InputValidationError, match="配对金额"):
        normalize_inputs(manifest)


def test_hidden_sheet_is_read_only_when_visible_formula_references_it(tmp_path):
    path = tmp_path / "隐藏引用.xlsx"
    workbook = Workbook()
    visible = workbook.active
    visible.title = "正表"
    visible["A1"] = "='审定表(2)'!A1"
    referenced = workbook.create_sheet("审定表(2)")
    referenced["A1"] = 1
    referenced.sheet_state = "hidden"
    ignored = workbook.create_sheet("无引用隐藏表")
    ignored["A1"] = 2
    ignored.sheet_state = "hidden"
    workbook.save(path)

    assert relevant_sheet_names(path) == ("正表", "审定表(2)")


def test_external_formula_does_not_activate_same_named_hidden_local_sheet(tmp_path):
    path = tmp_path / "外部引用.xlsx"
    workbook = Workbook()
    visible = workbook.active
    visible.title = "正表"
    visible["A1"] = "=[外部.xlsx]审定表(2)!A1"
    hidden = workbook.create_sheet("审定表(2)")
    hidden.sheet_state = "hidden"
    workbook.save(path)

    assert relevant_sheet_names(path) == ("正表",)


def test_optional_adjustments_and_trial_balance_auxiliary_fields_are_preserved(tmp_path):
    manifest = build_manifest(tmp_path)
    write_table(
        manifest.trial_balance_path,
        ["科目编码", "科目名称", "期初余额", "借方发生额", "贷方发生额", "期末余额", "辅助核算"],
        [["1002-01", "冻结银行存款", 100, 0, 0, 100, "受限资金"]],
    )
    book_adjustments = tmp_path / "账表调整.xlsx"
    audit_adjustments = tmp_path / "审计调整.xlsx"
    write_table(book_adjustments, ["调整编号", "报表项目", "金额", "性质"], [["B1", "应收账款", 10, "重分类"]])
    write_table(audit_adjustments, ["调整编号", "报表项目", "金额", "性质"], [["A1", "存货", -20, "审计调整"]])
    manifest = InputManifest(
        manifest.audited_balance_sheet_path,
        manifest.audited_income_statement_path,
        manifest.trial_balance_path,
        manifest.journal_pairs_path,
        manifest.prior_cashflow_path,
        manifest.display_unit,
        manifest.currency,
        manifest.performance_materiality_minor,
        book_adjustments,
        audit_adjustments,
    )

    bundle = normalize_inputs(manifest)

    assert bundle.trial_balance[0].original_fields["辅助核算"] == "受限资金"
    assert bundle.book_to_report_adjustments[0].adjustment_id == "B1"
    assert bundle.book_to_report_adjustments[0].amount_minor == 1_000
    assert bundle.book_to_report_adjustments[0].adjustment_type == "book_to_report"
    assert bundle.audit_adjustments[0].adjustment_id == "A1"
    assert bundle.audit_adjustments[0].amount_minor == -2_000
    assert bundle.audit_adjustments[0].adjustment_type == "audit"


def test_formula_without_cached_value_is_not_silently_read_as_zero(tmp_path):
    manifest = build_manifest(tmp_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["项目", "本期金额"])
    sheet.append(["营业收入", "=1+1"])
    workbook.save(manifest.audited_income_statement_path)

    with pytest.raises(InputFormulaCacheError, match=r"利润表.xlsx.*数据!B2.*公式没有缓存值"):
        normalize_inputs(manifest)
