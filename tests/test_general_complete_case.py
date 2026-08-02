from pathlib import Path

from dataclasses import replace

import openpyxl

from cashflow_main.contracts import EnterpriseType, InputManifest, RunStatus
from cashflow_main.pipeline import RunConfig, prepare_run


def save(path: Path, headers, rows=()):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


def build_complete_general_case(tmp_path: Path) -> RunConfig:
    bs = tmp_path / "审定资产负债表.xlsx"
    save(bs, ["项目", "期末数", "期初数"], [
        ("货币资金", 275, 120),
        ("应收账款", 80, 100),
        ("存货", 120, 100),
        ("固定资产", 100, 0),
        ("长期股权投资", 70, 20),
        ("应付账款", 90, 100),
        ("应付职工薪酬", 0, 80),
        ("应交税费", 0, 30),
        ("应付股利", 0, 20),
        ("短期借款", 150, 0),
        ("实收资本", 100, 0),
        ("资产总计", 670, 340),
    ])
    income = tmp_path / "审定利润表.xlsx"
    save(income, ["项目", "本期数"], [
        ("营业收入", 500),
        ("营业成本", 300),
        ("投资收益", 10),
        ("财务费用", 10),
        ("利润总额", 200),
    ])
    tb = tmp_path / "科目余额表.xlsx"
    save(tb, ["科目编码", "科目名称", "期初余额", "借方发生额", "贷方发生额", "期末余额", "资金性质"], [
        ("100201", "银行存款", 100, 850, 695, 255, "可随时支取"),
        ("100202", "银行存款-履约保证金", 20, 0, 0, 20, "受限"),
        ("1122", "应收账款", 100, 500, 520, 80, ""),
        ("1405", "存货", 100, 320, 300, 120, ""),
        ("1601", "固定资产", 0, 100, 0, 100, ""),
        ("1511", "长期股权投资", 20, 70, 20, 70, ""),
        ("2202", "应付账款", 100, 330, 320, 90, ""),
        ("2211", "应付职工薪酬", 80, 80, 0, 0, ""),
        ("2221", "应交税费", 30, 30, 0, 0, ""),
        ("2232", "应付股利", 20, 20, 0, 0, ""),
        ("2001", "短期借款", 0, 50, 200, 150, ""),
        ("4001", "实收资本", 0, 0, 100, 100, ""),
        ("4002", "股票发行费用", 0, 5, 0, 5, ""),
        ("6001", "营业收入", 0, 0, 500, 500, ""),
        ("6401", "营业成本", 0, 300, 0, 300, ""),
        ("6111", "投资收益", 0, 0, 10, 10, ""),
        ("6603", "财务费用", 0, 10, 0, 10, ""),
    ])
    journal = tmp_path / "已拆分一借一贷明细.xlsx"
    save(journal, ["借方科目", "贷方科目", "配对金额"], [
        ("应收账款", "营业收入", 500),
        ("银行存款", "应收账款", 520),
        ("存货", "应付账款", 320),
        ("营业成本", "存货", 300),
        ("应付账款", "银行存款", 330),
        ("应付职工薪酬", "银行存款", 80),
        ("应交税费", "银行存款", 30),
        ("固定资产", "银行存款", 100),
        ("长期股权投资", "银行存款", 70),
        ("银行存款", "长期股权投资", 20),
        ("银行存款", "投资收益", 10),
        ("银行存款", "短期借款", 200),
        ("短期借款", "银行存款", 50),
        ("财务费用", "银行存款", 10),
        ("银行存款", "实收资本", 100),
        ("股票发行费用", "银行存款", 5),
        ("应付股利", "银行存款", 20),
    ])
    prior = tmp_path / "上期审定现金流量表.xlsx"
    save(prior, ["项目", "本期数"])
    return RunConfig(
        InputManifest(bs, income, tb, journal, prior, "元", "人民币", 10_000),
        enterprise_type=EnterpriseType.GENERAL,
        entity_name="完整案例公司",
        period="2025年度",
    )


def test_complete_nonzero_general_case_runs_engine_and_reconciles_every_major_item(tmp_path):
    result = prepare_run(build_complete_general_case(tmp_path), tmp_path / "run")
    amounts = {item.item_id: item.amount_minor for item in result.calculation.items}

    assert result.ledger_reconciliation.is_reconciled
    assert result.adjustment_bridge.is_amount_reconciled
    assert result.status == RunStatus.VALIDATED
    assert result.validation_report.cash_change_difference_minor == 0
    assert amounts["CFO-01"] == 52_000
    assert amounts["CFO-04"] == 33_000
    assert amounts["CFO-05"] == 8_000
    assert amounts["CFO-06"] == 3_000
    assert amounts["CFO-NET"] == 8_000
    assert amounts["CFI-01"] == 2_000
    assert amounts["CFI-02"] == 1_000
    assert amounts["CFI-06"] == 10_000
    assert amounts["CFI-07"] == 7_000
    assert amounts["CFI-NET"] == -14_000
    assert amounts["CFF-01"] == 9_500
    assert amounts["CFF-02"] == 20_000
    assert amounts["CFF-04"] == 5_000
    assert amounts["CFF-05"] == 3_000
    assert amounts["CFF-06"] == 0
    assert amounts["CFF-NET"] == 21_500
    assert amounts["NET-CASH"] == 15_500
    assert amounts["OPENING-CASH"] == 10_000
    assert amounts["CLOSING-CASH"] == 25_500


def test_complete_case_has_no_cashflow_label_dependency(tmp_path):
    config = build_complete_general_case(tmp_path)
    result = prepare_run(config, tmp_path / "run")
    pair_details = [
        detail
        for item in result.calculation.items
        for component in item.components
        for detail in component.fact_details
        if detail.fact_id.startswith("JP:")
    ]
    assert pair_details
    assert all(detail.source_ids for detail in pair_details)


def test_material_conflicting_cashflow_label_is_visible_and_keeps_deterministic_amount(tmp_path):
    config = build_complete_general_case(tmp_path)
    workbook = openpyxl.load_workbook(config.manifest.journal_pairs_path)
    sheet = workbook.active
    sheet["D1"] = "现金流标签"
    sheet["D13"] = "investment_acquisition_cash"
    workbook.save(config.manifest.journal_pairs_path)

    result = prepare_run(config, tmp_path / "run")

    assert result.status == RunStatus.PROVISIONAL
    assert result.calculation.by_id["CFF-02"].amount_minor == 20_000
    conflict = next(
        case for case in result.decision_cases
        if case.decision_id.startswith("TAG_CONFLICT:")
    )
    assert conflict.human_review_required is True
    visible = result.review_context["decision_cases"]
    assert any(row["decision_id"] == conflict.decision_id for row in visible)


def test_material_unspecified_issue_cost_is_provisional_and_not_silently_netted(tmp_path):
    config = build_complete_general_case(tmp_path)
    for path in (config.manifest.trial_balance_path, config.manifest.journal_pairs_path):
        workbook = openpyxl.load_workbook(path)
        for row in workbook.active.iter_rows():
            for cell in row:
                if cell.value == "股票发行费用":
                    cell.value = "发行费用"
        workbook.save(path)
    manifest = replace(config.manifest, performance_materiality_minor=100)
    config = RunConfig(
        manifest,
        enterprise_type=config.enterprise_type,
        entity_name=config.entity_name,
        period=config.period,
    )

    result = prepare_run(config, tmp_path / "run")

    assert result.status == RunStatus.PROVISIONAL
    assert result.calculation.by_id["CFF-01"].amount_minor == 10_000
    assert result.calculation.by_id["CFF-06"].amount_minor == 500
    assert any(
        case.decision_id.startswith("FINANCING_ISSUE_COST:")
        and case.human_review_required
        for case in result.decision_cases
    )
