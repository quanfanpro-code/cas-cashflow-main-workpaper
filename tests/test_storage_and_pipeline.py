from pathlib import Path

import openpyxl
import pytest

from cashflow_main.contracts import EnterpriseType, InputManifest, RunStatus
from cashflow_main.input_adapter import InputFormulaCacheError
from cashflow_main.pipeline import RunConfig, finalize_run, get_status, load_run_artifacts, load_validation_report, prepare_run
from cashflow_main.storage import InputChangedError


def workbook(path: Path, headers, row=None):
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(headers)
    if row is not None: ws.append(row)
    wb.save(path)


def config(tmp_path):
    bs = tmp_path / "审定资产负债表.xlsx"; workbook(bs, ["项目", "期末数", "期初数"], ["货币资金", 0, 0])
    income = tmp_path / "审定利润表.xlsx"; workbook(income, ["项目", "本期数"], ["营业收入", 0])
    tb = tmp_path / "科目余额表.xlsx"; workbook(tb, ["科目编码", "科目名称", "期初余额", "借方发生额", "贷方发生额", "期末余额"], ["1002", "银行存款", 0, 0, 0, 0])
    journal = tmp_path / "一借一贷明细.xlsx"; workbook(journal, ["借方科目", "贷方科目", "配对金额"])
    prior = tmp_path / "上期现金流量表.xlsx"; workbook(prior, ["项目", "本期数"], ["销售商品、提供劳务收到的现金", 0])
    manifest = InputManifest(bs, income, tb, journal, prior, "元", "人民币", 100)
    return RunConfig(manifest, enterprise_type=EnterpriseType.GENERAL, entity_name="测试公司", period="2025年度")


def config_with_statement_difference(tmp_path, *, amount, materiality, adjustment=False):
    bs = tmp_path / "审定资产负债表.xlsx"
    workbook(
        bs,
        ["项目", "期末数", "期初数"],
        ["货币资金", 0, 0],
    )
    wb = openpyxl.load_workbook(bs)
    wb.active.append(["其他流动资产", amount, 0])
    wb.active.append(["流动资产合计", amount, 0])
    wb.save(bs)
    income = tmp_path / "审定利润表.xlsx"; workbook(income, ["项目", "本期数"], ["营业收入", 0])
    tb = tmp_path / "科目余额表.xlsx"; workbook(tb, ["科目编码", "科目名称", "期初余额", "借方发生额", "贷方发生额", "期末余额"], ["1002", "银行存款", 0, 0, 0, 0])
    journal = tmp_path / "一借一贷明细.xlsx"; workbook(journal, ["借方科目", "贷方科目", "配对金额"])
    prior = tmp_path / "上期现金流量表.xlsx"; workbook(prior, ["项目", "本期数"], ["销售商品、提供劳务收到的现金", 0])
    adjustment_path = None
    if adjustment:
        adjustment_path = tmp_path / "账表调整.xlsx"
        workbook(adjustment_path, ["调整编号", "报表项目", "调整金额", "调整性质"], ["BTR-1", "其他流动资产", amount, "非现金重分类"])
    manifest = InputManifest(
        bs, income, tb, journal, prior, "元", "人民币", materiality,
        book_to_report_adjustments_path=adjustment_path,
    )
    return RunConfig(manifest, enterprise_type=EnterpriseType.GENERAL)


def test_clean_run_reaches_validated_and_persists_status(tmp_path):
    result = prepare_run(config(tmp_path), tmp_path / "run")
    assert result.status == RunStatus.VALIDATED
    assert result.validation_report.is_clean
    assert get_status(result.run_dir).status == RunStatus.VALIDATED


def test_clean_run_persists_components_bridge_and_ledger_for_final_export(tmp_path):
    result = prepare_run(config(tmp_path), tmp_path / "run")
    calculation, bridge, ledger = load_run_artifacts(result.run_dir)
    first_component = calculation.items[0].components[0]
    assert first_component.operation
    assert first_component.source_scope
    assert calculation.items[0].verification_record_id
    assert bridge.rows
    assert ledger.is_reconciled
    assert load_validation_report(result.run_dir).is_clean


def test_input_change_invalidates_checkpoint(tmp_path):
    cfg = config(tmp_path)
    result = prepare_run(cfg, tmp_path / "run")
    with cfg.manifest.trial_balance_path.open("ab") as handle: handle.write(b"changed")
    with pytest.raises(InputChangedError, match="输入文件已变化"):
        finalize_run(result.run_dir, decisions=[])


def test_optional_adjustment_is_used_by_pipeline_and_control_total_is_ignored(tmp_path):
    result = prepare_run(
        config_with_statement_difference(tmp_path, amount=10, materiality=100, adjustment=True),
        tmp_path / "run",
    )
    rows = {row.report_item: row for row in result.adjustment_bridge.rows}
    assert rows["其他流动资产"].book_to_report_minor == 1_000
    assert rows["其他流动资产"].unexplained_minor == 0
    assert "流动资产合计" not in rows


def test_immaterial_unexplained_difference_is_auto_adopted_with_trace(tmp_path):
    result = prepare_run(
        config_with_statement_difference(tmp_path, amount=1, materiality=1_000),
        tmp_path / "run",
    )
    row = next(row for row in result.adjustment_bridge.rows if row.report_item == "其他流动资产")
    assert result.status == RunStatus.VALIDATED
    assert row.unexplained_minor == 0
    assert any(item.adjustment_id.startswith("AUTO:") for item in row.matched_adjustments)
    assert result.decision_cases[0].human_review_required is False


def test_material_cash_nature_conflict_stays_provisional(tmp_path):
    result = prepare_run(
        config_with_statement_difference(tmp_path, amount=10, materiality=1_000),
        tmp_path / "run",
    )
    assert result.status == RunStatus.PROVISIONAL
    assert result.decision_cases[0].strong_conflict is True
    assert result.decision_cases[0].human_review_required is True


def test_prepare_failure_persists_failed_state_and_chinese_error(tmp_path):
    cfg = config(tmp_path)
    workbook_with_formula = openpyxl.load_workbook(cfg.manifest.audited_income_statement_path)
    workbook_with_formula.active["B2"] = "=1+1"
    workbook_with_formula.save(cfg.manifest.audited_income_statement_path)

    with pytest.raises(InputFormulaCacheError):
        prepare_run(cfg, tmp_path / "run")

    state = __import__("json").loads((tmp_path / "run/state.json").read_text(encoding="utf-8-sig"))
    assert state["status"] == RunStatus.FAILED.value
    assert "公式没有缓存值" in state["error_message"]


def test_ledger_block_persists_differences_and_cannot_be_finalized(tmp_path):
    cfg = config(tmp_path)
    wb = openpyxl.load_workbook(cfg.manifest.trial_balance_path)
    wb.active["D2"] = 10
    wb.active["F2"] = 10
    wb.save(cfg.manifest.trial_balance_path)

    result = prepare_run(cfg, tmp_path / "run")

    assert result.status == RunStatus.BLOCKED
    evidence = __import__("json").loads(
        (result.run_dir / "ledger_reconciliation.json").read_text(encoding="utf-8-sig")
    )
    assert evidence["differences"]
    with pytest.raises(ValueError, match="不得输出最终"):
        finalize_run(result.run_dir, decisions=[])
