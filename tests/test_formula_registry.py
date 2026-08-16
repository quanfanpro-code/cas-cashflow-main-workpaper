import sys
import json
import subprocess
from pathlib import Path

from openpyxl import Workbook

SCRIPTS = Path(__file__).parents[1] / "scripts"
sys.path.insert(0, str(SCRIPTS))

from extract_formula_evidence import extract_xlsx  # noqa: E402
from validate_formula_registry import validate_formula_record  # noqa: E402


def valid_formula_record():
    return {
        "verification_id": "FV-G-001",
        "enterprise_type": "general",
        "cashflow_item_id": "CFO-01",
        "candidate_formula": {
            "source_locator": "third_workbook:现金流测试!D8",
            "components": [],
        },
        "evidence": {
            "knowledge_base": [{"locator": "知识库:#经营活动"}],
            "pdf_article": [{"locator": "chapter23.pdf:p10"}],
            "second_slides": [{"locator": "PPTX:p20"}],
            "second_workbook": [{"locator": "第二套:Sheet1!A1"}],
            "first_workbook": [{"locator": "第一套:正表!A1"}],
        },
        "issues": [],
        "corrected_formula": {
            "components": [],
            "gross_or_net": "gross",
            "noncash_exclusions": [],
            "special_adjustments": [],
            "restricted_cash_treatment": [],
        },
        "conclusion": "verified",
        "reviewed_at": "2026-08-02",
    }


def test_formula_record_requires_all_five_local_cross_checks():
    record = valid_formula_record()
    record["evidence"].pop("second_workbook")
    errors = validate_formula_record(record)
    assert any("second_workbook" in error for error in errors)


def test_title_only_period_mismatch_is_not_a_formula_error():
    record = valid_formula_record()
    record["issues"] = [
        {
            "type": "title_typo",
            "affects_formula": False,
            "resolution": "保留原始定位，不改变公式",
        }
    ]
    assert validate_formula_record(record) == []


def test_hidden_sheet_is_extracted_only_when_visible_formula_references_it(tmp_path):
    path = tmp_path / "来源.xlsx"
    workbook = Workbook()
    visible = workbook.active
    visible.title = "正表"
    visible["A1"] = "='审定表(2)'!A1"
    referenced = workbook.create_sheet("审定表(2)")
    referenced["A1"] = 1
    referenced.sheet_state = "hidden"
    ignored = workbook.create_sheet("无引用隐藏表")
    ignored["A1"] = 2
    ignored.sheet_state = "hidden"
    workbook.save(path)

    result = extract_xlsx(path, source_id="third_workbook")
    assert result.hidden_sheets_read == ("审定表(2)",)
    assert all("无引用隐藏表" not in item["locator"] for item in result.evidence)


def test_current_insurance_registry_uses_item_specific_gross_net_and_subtotal_basis():
    root = Path(__file__).parents[1]
    pack = json.loads(
        (root / "rules/insurance_2023_v1.json").read_text(encoding="utf-8-sig")
    )
    records = {
        record["cashflow_item_id"]: record
        for line in (root / "references/公式核验/insurance_2023_v1.jsonl")
        .read_text(encoding="utf-8-sig")
        .splitlines()
        if line.strip()
        for record in (json.loads(line),)
    }

    bases = set()
    for item in pack["items"]:
        record = records[item["item_id"]]
        if item.get("is_subtotal"):
            expected = "not_applicable"
        elif any(component["gross_or_net"] == "net" for component in item["components"]):
            expected = "net"
        elif any(component["gross_or_net"] == "gross" for component in item["components"]):
            expected = "gross"
        else:
            expected = "not_applicable"
        assert record["corrected_formula"]["gross_or_net"] == expected
        basis = record["netting_basis"]
        bases.add(basis)
        if expected == "gross":
            wording = basis + "".join(record["corrected_formula"]["special_adjustments"])
            assert "净增加额" not in wording
            assert "轧差" not in wording
        if expected == "net":
            assert "净额" in basis or "轧差" in basis

    assert len(bases) >= 4


def test_validate_script_all_rule_packs_checks_each_own_registry():
    script = Path(__file__).parents[1] / "scripts/validate_formula_registry.py"
    result = subprocess.run(
        [
            sys.executable,
            "-X",
            "utf8",
            str(script),
            "--all-rule-packs",
            "--require-zero-unresolved",
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
    )

    assert result.returncode == 0, result.stdout + result.stderr
    assert "missing_evidence=0" in result.stdout
    assert "orphan_rule=0" in result.stdout
