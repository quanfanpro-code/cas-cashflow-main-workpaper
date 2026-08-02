import json
from pathlib import Path

import openpyxl

from cashflow_main.adjustment_bridge import AdjustmentBridgeResult, AdjustmentBridgeRow, AdjustmentRecord
from cashflow_main.completeness_check import ValidationReport
from cashflow_main.contracts import EnterpriseType, InputManifest
from cashflow_main.fact_extraction import Fact, FactLedger
from cashflow_main.item_calculators import calculate_items
from cashflow_main.ledger_reconciliation import LedgerDifference, LedgerReconciliationResult
from cashflow_main.output import write_cashflow_statement
from cashflow_main.pipeline import RunConfig
from cashflow_main.rule_loader import load_rule_pack

ROOT = Path(__file__).parents[1]


def load_general_pack():
    registry = {
        json.loads(line)["verification_id"]
        for line in (ROOT / "references/公式核验/general_enterprise_v1.jsonl").read_text(encoding="utf-8-sig").splitlines()
        if line.strip()
    }
    return load_rule_pack(ROOT / "rules/general_enterprise_v1.json", registry)


def test_case3_style_workpaper_shows_components_sources_controls_and_links(tmp_path):
    facts = FactLedger.index((
        Fact("IS:营业收入", 10_000, ("statement", "is"), ("statement:IS:营业收入",), "K-REV", (("item_name", "营业收入"), ("period", "current"))),
        Fact("JP:1", 1_000, ("journal_pair", "short_term_borrowing_cash_received"), ("journal_pair:1",), "JP:1", (
            ("classification_evidence", ("借记现金、贷记短期借款",)),
            ("supplied_tags", ("investment_acquisition_cash",)),
            ("tag_conflicts", ("investment_acquisition_cash",)),
        )),
    ))
    calculation = calculate_items(load_general_pack(), facts)
    dummy = tmp_path / "input.xlsx"
    config = RunConfig(InputManifest(dummy, dummy, dummy, dummy, dummy, "元", "人民币", 100), EnterpriseType.GENERAL, "测试公司", "2025年度")
    orphan = AdjustmentRecord("A-ORPHAN", "不存在的项目", 300, "audit", "审计调整", ("调整表:第2行",))
    bridge = AdjustmentBridgeResult((
        AdjustmentBridgeRow("营业收入", 9_000, 500, 500, 10_000, 1_000, 0, (), ()),
    ), (orphan,), False)

    path = write_cashflow_statement(
        calculation, {}, config, ValidationReport((), (), (), 0), "最终",
        tmp_path / "现金流量表.xlsx", bridge=bridge,
        ledger_reconciliation=LedgerReconciliationResult(False, (
            LedgerDifference("银行存款", "debit", 100, 90, -10),
        )),
        review_context={
            "unmapped_accounts": [{"account_code": "9999", "account_name": "待映射科目", "closing_minor": 100}],
            "decision_cases": [{
                "decision_id": "D1", "human_review_required": True,
                "preferred_item_id": "A", "preferred_amount_minor": 100,
                "supporting_evidence": ["支持证据A"], "contrary_evidence": ["相反证据B"],
                "confirmation_status": "已确认",
            }],
                "restricted_cash": [
                    {"period": "期初", "account_name": "冻结存款", "amount_minor": 100, "source_ids": ["trial_balance:1002"]},
                    {"period": "期末", "account_name": "冻结存款", "amount_minor": 200, "source_ids": ["trial_balance:1002"]},
                ],
                "unallocated_cash": [{
                    "fact_id": "JP:9",
                    "debit_account_name": "其他应收款",
                    "credit_account_name": "银行存款",
                    "amount_minor": 300,
                    "source_ids": ["journal_pair:9"],
                    "evidence": ["摘要信息不足，未能分类"],
                }],
        },
    )
    workbook = openpyxl.load_workbook(path, data_only=False)
    assert workbook.sheetnames == ["现金流量表", "计算验证底稿"]
    statement = workbook["现金流量表"]
    workpaper = workbook["计算验证底稿"]
    values = [cell.value for row in workpaper.iter_rows() for cell in row if cell.value not in (None, "")]
    for expected in (
        "账表调整桥", "账面数", "账表调整", "审计调整", "审定数", "未解释差额",
        "组成编号", "组成或来源项目", "原始金额", "加减方向", "计入金额", "来源定位",
        "非现金剔除", "特殊事项调整", "受限资金处理", "规则计算结果", "正表列示金额", "验证差异", "核对结论",
        "statement:IS:营业收入", "CFO-01-C01", "未命中", "账务核对", "不一致（1项）",
        "未映射科目", "待映射科目", "自动判断与待确认事项", "D1", "支持证据A", "相反证据B", "已确认",
            "现金范围剔除明细", "冻结存款", "期初", "期末", "账务核对差异", "银行存款",
            "未分类现金明细", "JP:9", "其他应收款", "摘要信息不足，未能分类",
        "孤立调整", "A-ORPHAN", "调整表:第2行", "自动分类依据", "外部标签", "标签冲突",
        "借记现金、贷记短期借款", "investment_acquisition_cash",
    ):
        assert expected in values
    assert any(cell.hyperlink and cell.hyperlink.location.startswith("'计算验证底稿'") for cell in statement["A"])
    assert any(cell.hyperlink and cell.hyperlink.location.startswith("'现金流量表'") for cell in workpaper["A"])
    formulas = [cell.value for row in workpaper.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")]
    assert any("SUM(" in formula for formula in formulas)
    assert any("'现金流量表'!" in formula for formula in formulas)
    assert not workbook._external_links
    assert all(sheet.sheet_state == "visible" for sheet in workbook.worksheets)


def test_final_statement_rejects_unconfirmed_human_review_case(tmp_path):
    calculation = calculate_items(load_general_pack(), FactLedger.index(()))
    dummy = tmp_path / "input.xlsx"
    config = RunConfig(InputManifest(dummy, dummy, dummy, dummy, dummy, "元", "人民币", 100), EnterpriseType.GENERAL)

    import pytest

    with pytest.raises(ValueError, match="人工确认"):
        write_cashflow_statement(
            calculation, {}, config, ValidationReport((), (), (), 0), "最终",
            tmp_path / "不应生成.xlsx",
            review_context={"decision_cases": [{
                "decision_id": "D1", "human_review_required": True,
                "confirmation_status": "待确认",
            }]},
        )
