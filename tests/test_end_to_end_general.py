import json
from pathlib import Path

import openpyxl

from cashflow_main.__main__ import main


def save(path: Path, headers, rows=()):
    wb = openpyxl.Workbook(); ws = wb.active; ws.append(headers)
    for row in rows: ws.append(row)
    wb.save(path)


def test_general_manifest_runs_to_statement_and_verification_workpaper(tmp_path, capsys):
    bs = tmp_path / "资产负债表.xlsx"; save(bs, ["项目", "期末数", "期初数"], [("货币资金", 0, 0)])
    income = tmp_path / "利润表.xlsx"; save(income, ["项目", "本期数"], [("营业收入", 0)])
    tb = tmp_path / "余额表.xlsx"; save(tb, ["科目编码", "科目名称", "期初余额", "借方发生额", "贷方发生额", "期末余额"], [("1002", "银行存款", 0, 0, 0, 0)])
    journal = tmp_path / "一借一贷.xlsx"; save(journal, ["借方科目", "贷方科目", "配对金额"])
    prior = tmp_path / "上期现流.xlsx"; save(prior, ["项目", "本期数"], [("销售商品、提供劳务收到的现金", 88)])
    manifest = tmp_path / "manifest.json"
    manifest.write_text(json.dumps({
        "audited_balance_sheet_path": str(bs), "audited_income_statement_path": str(income),
        "trial_balance_path": str(tb), "journal_pairs_path": str(journal), "prior_cashflow_path": str(prior),
        "display_unit": "元", "currency": "人民币", "performance_materiality_minor": 100,
        "enterprise_type": "general", "entity_name": "端到端测试公司", "period": "2025年度", "output_dir": str(tmp_path),
    }, ensure_ascii=False), encoding="utf-8-sig")
    assert main(["prepare", "--manifest", str(manifest)]) == 0
    result = json.loads(capsys.readouterr().out)
    assert result["status"] == "validated"
    wb = openpyxl.load_workbook(result["output"], data_only=False)
    assert wb.sheetnames == ["现金流量表", "计算验证底稿"]
    assert wb.active["D6"].value == 88
