from pathlib import Path

import json

import openpyxl

from cashflow_main.__main__ import build_parser, main

ROOT = Path(__file__).parents[1]


def test_cli_exposes_only_stable_commands():
    parser = build_parser()
    choices = parser._subparsers._group_actions[0].choices
    assert set(choices) == {"prepare", "finalize", "status"}


def test_skill_guidance_names_fixed_inputs_and_final_gate():
    text = (ROOT / "SKILL.md").read_text(encoding="utf-8-sig")
    for value in (
        "一借一贷明细", "实际执行重要性水平", "暂编正表", "不得标记为最终正表",
        "直接法", "不处理", "不重新配对", "借贷科目关系", "计算验证底稿", "两张可见工作表",
        "内部复核公式", "上期数只搬取、不复算",
    ):
        assert value in text


def test_launcher_does_not_depend_on_current_directory():
    text = (ROOT / "scripts/engine_launcher.py").read_text(encoding="utf-8-sig")
    assert "__file__" in text
    assert "cashflow_main.__main__" in text


def test_windows_selector_collects_entity_name_and_period():
    text = (ROOT / "scripts/select_paths.py").read_text(encoding="utf-8-sig")
    assert '"entity_name"' in text
    assert '"period"' in text


def _workbook(path: Path, headers, rows):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def test_prepare_command_exports_bridge_and_ledger_into_visible_workpaper(tmp_path, capsys):
    balance_sheet = tmp_path / "审定资产负债表.xlsx"
    income_statement = tmp_path / "审定利润表.xlsx"
    trial_balance = tmp_path / "科目余额表.xlsx"
    journal_pairs = tmp_path / "一借一贷明细.xlsx"
    prior_cashflow = tmp_path / "上期现金流量表.xlsx"
    _workbook(balance_sheet, ["项目", "期末数", "期初数"], [["货币资金", 1, 0]])
    _workbook(income_statement, ["项目", "本期数"], [["营业收入", 1]])
    _workbook(trial_balance, ["科目编码", "科目名称", "期初余额", "借方发生额", "贷方发生额", "期末余额"], [
        ["1002", "银行存款", 0, 1, 0, 1],
        ["6001", "主营业务收入", 0, 0, 1, 1],
    ])
    _workbook(journal_pairs, ["借方科目", "贷方科目", "配对金额"], [["银行存款", "主营业务收入", 1]])
    _workbook(prior_cashflow, ["项目", "本期数"], [["销售商品、提供劳务收到的现金", 0]])
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps({
        "audited_balance_sheet_path": str(balance_sheet),
        "audited_income_statement_path": str(income_statement),
        "trial_balance_path": str(trial_balance),
        "journal_pairs_path": str(journal_pairs),
        "prior_cashflow_path": str(prior_cashflow),
        "display_unit": "元",
        "currency": "人民币",
        "performance_materiality_minor": 100,
        "enterprise_type": "general",
        "entity_name": "测试公司",
        "period": "2025年度",
    }, ensure_ascii=False), encoding="utf-8-sig")
    run_dir = tmp_path / "run"

    assert main(["prepare", "--manifest", str(manifest_path), "--run-dir", str(run_dir)]) == 0
    result = json.loads(capsys.readouterr().out)
    workbook = openpyxl.load_workbook(result["output"], data_only=False)
    rows = list(workbook["计算验证底稿"].iter_rows(values_only=True))
    assert any(row[0] == "营业收入" and row[7] == "已解释" for row in rows)
    assert any("已勾稽" in row for row in rows)
