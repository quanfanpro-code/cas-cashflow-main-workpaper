"""读取并标准化审定报表、科目余额表和一借一贷明细。"""

import re
from pathlib import Path

from openpyxl import load_workbook

from .contracts import (
    AccountBalance,
    InputManifest,
    JournalPair,
    NormalizedAdjustment,
    NormalizedInputBundle,
    StatementLine,
)
from .money import to_minor_units


class InputValidationError(ValueError):
    """输入资料缺失或格式无法可靠识别。"""


class InputFormulaCacheError(InputValidationError):
    """输入工作簿含公式，但没有可读取的已保存结果。"""


ALIASES = {
    "item_name": {"项目", "项目名称", "报表项目"},
    "current_amount": {"本期金额", "本期数", "期末余额", "期末数", "年末余额"},
    "prior_amount": {"上期金额", "上期数", "期初余额", "期初数", "年初余额"},
    "account_code": {"科目编码", "科目代码"},
    "account_name": {"科目名称", "会计科目"},
    "opening_balance": {"期初余额", "年初余额"},
    "debit_turnover": {"借方发生额", "本期借方"},
    "credit_turnover": {"贷方发生额", "本期贷方"},
    "closing_balance": {"期末余额", "年末余额"},
    "debit_account_name": {"借方科目", "借方科目名称"},
    "credit_account_name": {"贷方科目", "贷方科目名称"},
    "paired_amount": {"配对金额", "匹配金额"},
    "adjustment_id": {"调整编号", "编号", "序号"},
    "report_item": {"报表项目", "项目", "项目名称"},
    "adjustment_amount": {"调整金额", "金额"},
    "adjustment_nature": {"调整性质", "性质", "业务性质"},
}
FIELD_NAMES = {
    "item_name": "项目",
    "current_amount": "本期或期末金额",
    "prior_amount": "上期或期初金额",
    "account_code": "科目编码",
    "account_name": "科目名称",
    "opening_balance": "期初余额",
    "debit_turnover": "借方发生额",
    "credit_turnover": "贷方发生额",
    "closing_balance": "期末余额",
    "debit_account_name": "借方科目",
    "credit_account_name": "贷方科目",
    "paired_amount": "配对金额",
    "adjustment_id": "调整编号",
    "report_item": "报表项目",
    "adjustment_amount": "调整金额",
    "adjustment_nature": "调整性质",
}
FORMULA_SHEET_REFERENCE = re.compile(
    r"'([^']+)'!|([A-Za-z0-9_\u4e00-\u9fff .()（）-]+)!"
)


def _header(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip()


def relevant_sheet_names(path: Path) -> tuple[str, ...]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    visible = [sheet.title for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
    referenced: list[str] = []
    for sheet_name in visible:
        for row in workbook[sheet_name].iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                for match in FORMULA_SHEET_REFERENCE.finditer(cell.value):
                    if match.start() > 0 and cell.value[match.start() - 1] == "]":
                        continue
                    name = (match.group(1) or match.group(2)).strip()
                    if (
                        name in workbook.sheetnames
                        and workbook[name].sheet_state != "visible"
                        and name not in referenced
                    ):
                        referenced.append(name)
    workbook.close()
    return tuple(visible + referenced)


def _canonical_columns(headers: list[object]) -> dict[str, int]:
    normalized = [_header(value) for value in headers]
    found: dict[str, int] = {}
    for canonical, aliases in ALIASES.items():
        for index, value in enumerate(normalized):
            if value in aliases:
                found[canonical] = index
                break
    return found


def _table_rows(
    path: Path,
    required: set[str],
) -> tuple[list[object], list[dict[str, object]], dict[str, int]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    formula_workbook = load_workbook(path, data_only=False, read_only=True)
    best: tuple[int, object, int, dict[str, int]] | None = None
    relevant = relevant_sheet_names(path)
    for sheet_order, sheet_name in enumerate(relevant):
        sheet = workbook[sheet_name]
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            columns = _canonical_columns(list(row))
            score = len(required & columns.keys())
            candidate = (score, -sheet_order, -row_number, columns)
            if best is None or candidate[:3] > best[:3]:
                best = (score, sheet, row_number, columns)
            if required <= columns.keys():
                headers = list(row)
                records = []
                formula_sheet = formula_workbook[sheet_name]
                for data_row_number, values in enumerate(sheet.iter_rows(
                    min_row=row_number + 1,
                    values_only=True,
                ), start=row_number + 1):
                    if not any(value not in (None, "") for value in values):
                        continue
                    formula_values = next(
                        formula_sheet.iter_rows(
                            min_row=data_row_number,
                            max_row=data_row_number,
                        )
                    )
                    for column_number, formula_cell in enumerate(formula_values, 1):
                        cached = values[column_number - 1] if column_number <= len(values) else None
                        if formula_cell.data_type == "f" and cached is None:
                            workbook.close()
                            formula_workbook.close()
                            raise InputFormulaCacheError(
                                f"{path.name}的{sheet_name}!{formula_cell.coordinate}公式没有缓存值，请先用Excel打开并保存"
                            )
                    records.append(
                        {
                            str(headers[index]): value
                            for index, value in enumerate(values)
                            if index < len(headers)
                        }
                    )
                workbook.close()
                formula_workbook.close()
                return headers, records, columns
    workbook.close()
    formula_workbook.close()
    missing_keys = required - (best[3].keys() if best else set())
    missing = "、".join(FIELD_NAMES[key] for key in sorted(missing_keys))
    raise InputValidationError(f"{path.name}缺少必要字段：{missing}")


def _value(record: dict[str, object], headers: list[object], index: int) -> object:
    return record.get(str(headers[index]))


def _read_statement(
    path: Path,
    display_unit: str,
    *,
    require_prior: bool,
) -> tuple[StatementLine, ...]:
    required = {"item_name", "current_amount"}
    if require_prior:
        required.add("prior_amount")
    headers, records, columns = _table_rows(path, required)
    lines = []
    for record in records:
        item_name = str(_value(record, headers, columns["item_name"]) or "").strip()
        if not item_name:
            continue
        current = _value(record, headers, columns["current_amount"])
        prior = (
            _value(record, headers, columns["prior_amount"])
            if "prior_amount" in columns
            else None
        )
        lines.append(
            StatementLine(
                item_name=item_name,
                current_minor=to_minor_units(current or 0, display_unit),
                prior_minor=(
                    to_minor_units(prior or 0, display_unit)
                    if prior is not None
                    else None
                ),
            )
        )
    return tuple(lines)


def _read_trial_balance(
    path: Path,
    display_unit: str,
) -> tuple[AccountBalance, ...]:
    required = {
        "account_code",
        "account_name",
        "opening_balance",
        "debit_turnover",
        "credit_turnover",
        "closing_balance",
    }
    headers, records, columns = _table_rows(path, required)
    rows = []
    for record in records:
        account_name = str(_value(record, headers, columns["account_name"]) or "").strip()
        if not account_name:
            continue
        rows.append(
            AccountBalance(
                account_code=str(
                    _value(record, headers, columns["account_code"]) or ""
                ).strip(),
                account_name=account_name,
                opening_balance_minor=to_minor_units(
                    _value(record, headers, columns["opening_balance"]) or 0,
                    display_unit,
                ),
                debit_turnover_minor=to_minor_units(
                    _value(record, headers, columns["debit_turnover"]) or 0,
                    display_unit,
                ),
                credit_turnover_minor=to_minor_units(
                    _value(record, headers, columns["credit_turnover"]) or 0,
                    display_unit,
                ),
                closing_balance_minor=to_minor_units(
                    _value(record, headers, columns["closing_balance"]) or 0,
                    display_unit,
                ),
                original_fields=dict(record),
            )
        )
    return tuple(rows)


def _read_adjustments(
    path: Path,
    display_unit: str,
    adjustment_type: str,
) -> tuple[NormalizedAdjustment, ...]:
    required = {"adjustment_id", "report_item", "adjustment_amount"}
    headers, records, columns = _table_rows(path, required)
    rows = []
    for index, record in enumerate(records, 1):
        adjustment_id = str(_value(record, headers, columns["adjustment_id"]) or "").strip()
        report_item = str(_value(record, headers, columns["report_item"]) or "").strip()
        if not adjustment_id or not report_item:
            continue
        nature = (
            str(_value(record, headers, columns["adjustment_nature"]) or "").strip()
            if "adjustment_nature" in columns
            else ""
        )
        rows.append(NormalizedAdjustment(
            adjustment_id=adjustment_id,
            report_item=report_item,
            amount_minor=to_minor_units(
                _value(record, headers, columns["adjustment_amount"]) or 0,
                display_unit,
            ),
            adjustment_type=adjustment_type,
            nature=nature or None,
            source_ids=(f"{path.name}:row:{index}",),
        ))
    return tuple(rows)


def _read_journal_pairs(
    path: Path,
    display_unit: str,
) -> tuple[JournalPair, ...]:
    required = {"debit_account_name", "credit_account_name", "paired_amount"}
    headers, records, columns = _table_rows(path, required)
    rows = []
    for record in records:
        debit = str(
            _value(record, headers, columns["debit_account_name"]) or ""
        ).strip()
        credit = str(
            _value(record, headers, columns["credit_account_name"]) or ""
        ).strip()
        if not debit and not credit:
            continue
        rows.append(
            JournalPair(
                debit_account_name=debit,
                credit_account_name=credit,
                amount_minor=to_minor_units(
                    _value(record, headers, columns["paired_amount"]) or 0,
                    display_unit,
                ),
                original_fields=dict(record),
            )
        )
    return tuple(rows)


def normalize_inputs(manifest: InputManifest) -> NormalizedInputBundle:
    if manifest.performance_materiality_minor <= 0:
        raise InputValidationError("实际执行重要性水平必须大于0")
    for name, path in manifest.input_paths().items():
        if not path.is_file():
            raise InputValidationError(f"固定输入不存在：{name}={path}")
    return NormalizedInputBundle(
        audited_balance_sheet=_read_statement(
            manifest.audited_balance_sheet_path,
            manifest.display_unit,
            require_prior=True,
        ),
        audited_income_statement=_read_statement(
            manifest.audited_income_statement_path,
            manifest.display_unit,
            require_prior=False,
        ),
        trial_balance=_read_trial_balance(
            manifest.trial_balance_path,
            manifest.display_unit,
        ),
        journal_pairs=_read_journal_pairs(
            manifest.journal_pairs_path,
            manifest.display_unit,
        ),
        prior_cashflow=_read_statement(
            manifest.prior_cashflow_path,
            manifest.display_unit,
            require_prior=False,
        ),
        book_to_report_adjustments=(
            _read_adjustments(
                manifest.book_to_report_adjustments_path,
                manifest.display_unit,
                "book_to_report",
            )
            if manifest.book_to_report_adjustments_path
            else ()
        ),
        audit_adjustments=(
            _read_adjustments(
                manifest.audit_adjustments_path,
                manifest.display_unit,
                "audit",
            )
            if manifest.audit_adjustments_path
            else ()
        ),
    )
