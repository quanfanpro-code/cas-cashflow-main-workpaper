"""输出数值正表和一张可见、可追溯、可反算的计算验证底稿。"""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.hyperlink import Hyperlink

from .completeness_check import ValidationReport
from .item_calculators import CalculationResult
from .money import from_minor_units
from .pipeline import RunConfig


_OPERATION_NAMES = {
    "statement_value": "审定报表取数",
    "balance_change": "余额变动",
    "debit_turnover": "借方发生额",
    "credit_turnover": "贷方发生额",
    "paired_turnover": "序时账配对金额",
    "adjustment_amount": "调整事项",
    "fact_amount": "补充事实取数",
    "net_fact_amount": "正负现金事实净额",
    "subtotal": "已计算项目汇总",
    "cash_equivalent_balance": "现金及现金等价物余额",
}

_SOURCE_NAMES = {
    "audited_statements": "审定报表",
    "trial_balance": "科目余额表",
    "fact_ledger": "已提取业务事实",
    "calculated_items": "本表已计算项目",
}

_NUMBER_FORMAT = "#,##0.00;[Red]-#,##0.00"


def _display_amount(amount_minor: int, display_unit: str) -> float:
    return float(from_minor_units(amount_minor, display_unit))


def _join(values: tuple[str, ...]) -> str:
    return "、".join(values) if values else "无"


def _selector_label(selector: dict[str, object], operation: str) -> str:
    if selector.get("item_name_groups"):
        return "；否则".join(
            "、".join(str(value) for value in group)
            for group in selector["item_name_groups"]
        )
    if selector.get("item_names"):
        return "、".join(str(value) for value in selector["item_names"])
    if selector.get("account_groups"):
        groups = "、".join(str(value) for value in selector["account_groups"])
        direction = {
            "opening_minus_closing": "期初－期末",
            "closing_minus_opening": "期末－期初",
        }.get(str(selector.get("direction")), "")
        return f"{groups}（{direction}）" if direction else groups
    if selector.get("account_names"):
        return "、".join(str(value) for value in selector["account_names"])
    if operation == "subtotal":
        parts = [str(value) for value in selector.get("item_ids", [])]
        parts.extend(f"减：{value}" for value in selector.get("subtract_item_ids", []))
        return "；".join(parts) or "本项目组成汇总"
    if operation == "cash_equivalent_balance":
        return "期初现金及现金等价物余额" if selector.get("period") == "opening" else "期末现金及现金等价物余额"
    if selector.get("positive_tags_any") or selector.get("negative_tags_any"):
        positive = "、".join(str(value) for value in selector.get("positive_tags_any", ()))
        negative = "、".join(str(value) for value in selector.get("negative_tags_any", ()))
        return f"正向：{positive or '无'}；负向：{negative or '无'}"
    if selector.get("tags_any"):
        return "、".join(str(value) for value in selector["tags_any"])
    return "规则指定的相关金额"


def _direction(component) -> str:
    direction_sign = -1 if component.selector.get("direction") == "opening_minus_closing" else 1
    return "加" if component.sign * direction_sign == 1 else "减"


def _set_internal_link(cell, location: str) -> None:
    cell.hyperlink = Hyperlink(ref=cell.coordinate, location=location)
    cell.style = "Hyperlink"


def _prior_map(prior_cashflow, display_unit: str) -> dict[str, float]:
    if isinstance(prior_cashflow, dict):
        return {str(key): float(value) for key, value in prior_cashflow.items()}
    return {
        line.item_name: float(from_minor_units(line.current_minor, display_unit))
        for line in prior_cashflow
    }


def _write_prior_matching(ws, start_row: int, unmatched: tuple[str, ...]) -> int:
    ws.cell(start_row, 1, "上期项目匹配")
    ws.cell(start_row, 1).font = Font(name="宋体", bold=True, size=12)
    if not unmatched:
        ws.cell(start_row + 1, 1, "全部项目已匹配")
        return start_row + 2
    for row_number, item_name in enumerate(unmatched, start_row + 1):
        ws.cell(row_number, 1, item_name)
        ws.cell(row_number, 2, "未匹配，上期数暂列0")
    return start_row + len(unmatched) + 1


def _write_review_context(ws, start_row: int, review_context: dict[str, object], display_unit: str) -> int:
    current = start_row
    sections = (
        ("未映射科目", "unmapped_accounts", ("科目编码", "科目名称", "期末余额")),
        ("自动判断与待确认事项", "decision_cases", ("事项编号", "是否人工复核", "首选判断", "候选判断", "影响金额", "支持证据", "相反证据", "确认状态")),
        ("受限资金明细", "restricted_cash", ("期间", "科目名称", "剔除金额", "来源")),
        ("账务核对差异", "ledger_differences", ("科目名称", "方向", "余额表金额", "序时账金额", "差异", "差异类型")),
    )
    for title, key, headers in sections:
        ws.cell(current, 1, title)
        ws.cell(current, 1).font = Font(name="宋体", bold=True, size=12)
        rows = tuple(review_context.get(key, ()))
        if not rows:
            ws.cell(current + 1, 1, "无")
            current += 3
            continue
        for column, header in enumerate(headers, 1):
            ws.cell(current + 1, column, header).font = Font(bold=True)
        for row_number, row in enumerate(rows, current + 2):
            if key == "unmapped_accounts":
                values = (
                    row.get("account_code", ""),
                    row.get("account_name", ""),
                    _display_amount(int(row.get("closing_minor", 0)), display_unit),
                )
            elif key == "decision_cases":
                values = (
                    row.get("decision_id", ""),
                    "是" if row.get("human_review_required") else "否",
                    row.get("preferred_item_id", ""),
                    "、".join(str(value) for value in row.get("candidate_item_ids", [])) or "无",
                    _display_amount(int(row.get("preferred_amount_minor", 0)), display_unit),
                    "、".join(str(value) for value in row.get("supporting_evidence", [])) or "无",
                    "、".join(str(value) for value in row.get("contrary_evidence", [])) or "无",
                    row.get("confirmation_status", "未记录"),
                )
            elif key == "restricted_cash":
                values = (
                    row.get("period", ""),
                    row.get("account_name", ""),
                    _display_amount(int(row.get("amount_minor", 0)), display_unit),
                    "、".join(str(value) for value in row.get("source_ids", [])),
                )
            else:
                values = (
                    row.get("account_name", ""),
                    row.get("side", ""),
                    _display_amount(int(row.get("expected_minor", 0)), display_unit),
                    _display_amount(int(row.get("actual_minor", 0)), display_unit),
                    _display_amount(int(row.get("difference_minor", 0)), display_unit),
                    row.get("kind", ""),
                )
            for column, value in enumerate(values, 1):
                ws.cell(row_number, column, value)
        current += len(rows) + 3
    return current


def _write_bridge(ws, start_row: int, bridge, display_unit: str) -> int:
    ws.cell(start_row, 1, "账表调整桥")
    ws.cell(start_row, 1).font = Font(name="宋体", bold=True, size=12)
    headers = (
        "报表项目", "账面数", "账表调整", "审计调整", "审定数", "总差额", "未解释差额", "状态",
        "匹配调整及性质", "调整来源",
    )
    header_row = start_row + 1
    for column, value in enumerate(headers, 1):
        cell = ws.cell(header_row, column, value)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")
    rows = tuple(getattr(bridge, "rows", ())) if bridge is not None else ()
    if not rows and not getattr(bridge, "orphan_adjustments", ()):
        ws.cell(header_row + 1, 1, "未提供账表调整桥明细")
        ws.cell(header_row + 1, 8, "未提供")
        return header_row + 2
    for row_number, row in enumerate(rows, header_row + 1):
        values = (
            row.report_item,
            _display_amount(row.book_minor, display_unit),
            _display_amount(row.book_to_report_minor, display_unit),
            _display_amount(row.audit_adjustment_minor, display_unit),
            _display_amount(row.audited_minor, display_unit),
            _display_amount(row.total_difference_minor, display_unit),
            _display_amount(row.unexplained_minor, display_unit),
            "已解释" if row.unexplained_minor == 0 else "待确认",
            "、".join(
                f"{item.adjustment_id}（{item.nature or item.adjustment_type}）"
                for item in row.matched_adjustments
            ) or "无",
            "、".join(
                source
                for item in row.matched_adjustments
                for source in item.source_ids
            ) or "无",
        )
        for column, value in enumerate(values, 1):
            ws.cell(row_number, column, value)
        for column in range(2, 8):
            ws.cell(row_number, column).number_format = _NUMBER_FORMAT
    next_row = header_row + len(rows) + 2
    orphans = tuple(getattr(bridge, "orphan_adjustments", ())) if bridge is not None else ()
    ws.cell(next_row, 1, "孤立调整")
    ws.cell(next_row, 1).font = Font(name="宋体", bold=True, size=12)
    if not orphans:
        ws.cell(next_row + 1, 1, "无")
        return next_row + 2
    orphan_headers = ("调整编号", "报表项目", "金额", "调整类型", "性质", "来源")
    for column, value in enumerate(orphan_headers, 1):
        ws.cell(next_row + 1, column, value).font = Font(bold=True)
    for row_number, item in enumerate(orphans, next_row + 2):
        values = (
            item.adjustment_id,
            item.report_item,
            _display_amount(item.amount_minor, display_unit),
            item.adjustment_type,
            item.nature or "未说明",
            "、".join(item.source_ids) or "无",
        )
        for column, value in enumerate(values, 1):
            ws.cell(row_number, column, value)
        ws.cell(row_number, 3).number_format = _NUMBER_FORMAT
    return next_row + len(orphans) + 2


def _write_verification_workpaper(
    wb: Workbook,
    calculation: CalculationResult,
    config: RunConfig,
    validation: ValidationReport,
    statement_kind: str,
    statement_rows: dict[str, int],
    bridge,
    ledger_reconciliation,
    prior_unmatched: tuple[str, ...],
    review_context: dict[str, object],
) -> None:
    ws = wb.create_sheet("计算验证底稿")
    ws.merge_cells("A1:O1")
    ws["A1"] = "现金流量表计算验证底稿"
    ws["A1"].font = Font(name="宋体", size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:O2")
    ws["A2"] = f"{config.entity_name}　{config.period}　{statement_kind}"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A4"] = "编制状态"
    ws["B4"] = statement_kind
    ws["D4"] = "完整性检查"
    ws["E4"] = "通过" if validation.is_clean else "存在未解决事项"
    ws["G4"] = "现金变动验证差异"
    ws["H4"] = _display_amount(validation.cash_change_difference_minor, config.manifest.display_unit)
    ws["H4"].number_format = _NUMBER_FORMAT
    ws["J4"] = "账务核对"
    if ledger_reconciliation is None:
        ws["K4"] = "未提供"
    else:
        ws["K4"] = "已勾稽" if ledger_reconciliation.is_reconciled else f"不一致（{len(ledger_reconciliation.differences)}项）"
    ws["A5"] = "说明"
    ws["B5"] = "正表金额由程序计算后固化；本底稿展示组成、来源、调整规则，并用表内公式反算正表。"
    ws.merge_cells("B5:O5")

    next_row = _write_prior_matching(ws, 7, prior_unmatched) + 2
    visible_review_context = dict(review_context)
    if ledger_reconciliation is not None and "ledger_differences" not in visible_review_context:
        visible_review_context["ledger_differences"] = [
            {
                "account_name": item.account_name,
                "side": item.side,
                "expected_minor": item.expected_minor,
                "actual_minor": item.actual_minor,
                "difference_minor": item.difference_minor,
                "kind": item.kind,
            }
            for item in ledger_reconciliation.differences
        ]
    next_row = _write_review_context(ws, next_row, visible_review_context, config.manifest.display_unit) + 1
    next_row = _write_bridge(ws, next_row, bridge, config.manifest.display_unit) + 2
    workpaper_rows: dict[str, int] = {}
    headers = (
        "组成编号", "组成或来源项目", "取数方式", "原始金额", "加减方向", "计入金额",
        "来源定位", "规则与核验", "非现金剔除", "特殊事项调整", "受限资金处理",
        "自动分类依据", "外部标签", "标签冲突", "状态",
    )
    for item in sorted(calculation.items, key=lambda value: value.display_order):
        title_row = next_row
        workpaper_rows[item.item_id] = title_row
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=15)
        title = ws.cell(title_row, 1, f"{item.item_id}　{item.name}")
        title.font = Font(name="宋体", bold=True, color="FFFFFF")
        title.fill = PatternFill("solid", fgColor="4472C4")
        statement_row = statement_rows[item.item_id]
        _set_internal_link(title, f"'现金流量表'!A{statement_row}")
        header_row = title_row + 1
        for column, value in enumerate(headers, 1):
            cell = ws.cell(header_row, column, value)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        detail_row = header_row + 1
        for component in item.components:
            facts = component.fact_details or (None,)
            for fact in facts:
                if fact is None:
                    raw_minor = component.amount_minor if component.operation == "subtotal" else 0
                    applied_minor = component.amount_minor
                    label = component.selector_label or _selector_label(component.selector, component.operation)
                    sources = _SOURCE_NAMES.get(component.source_scope, component.source_scope or "无")
                    status = "汇总" if component.operation == "subtotal" else "未命中"
                else:
                    raw_minor = fact.raw_amount_minor
                    applied_minor = fact.applied_amount_minor
                    label = fact.fact_label
                    sources = "、".join(fact.source_ids) or _SOURCE_NAMES.get(component.source_scope, component.source_scope or "无")
                    status = (
                        "净额方向未列示"
                        if component.operation == "net_fact_amount"
                        and fact.applied_amount_minor == 0
                        and fact.raw_amount_minor != 0
                        else "已取数"
                    )
                classification_evidence = _join(fact.classification_evidence) if fact is not None else "无"
                supplied_tags = _join(fact.supplied_tags) if fact is not None else "无"
                tag_conflicts = _join(fact.tag_conflicts) if fact is not None else "无"
                direction = (
                    "加" if applied_minor >= 0 else "减"
                    if component.operation == "net_fact_amount"
                    else _direction(component)
                )
                values = (
                    component.rule_component_id,
                    label,
                    _OPERATION_NAMES.get(component.operation, component.operation or "未记录"),
                    _display_amount(raw_minor, config.manifest.display_unit),
                    direction,
                    _display_amount(applied_minor, config.manifest.display_unit),
                    sources,
                    f"{item.verification_record_id or '未记录核验编号'}；{component.gross_or_net}",
                    _join(component.noncash_exclusions),
                    _join(component.special_adjustments),
                    _join(component.restricted_cash_treatment),
                    classification_evidence,
                    supplied_tags,
                    tag_conflicts,
                    status,
                )
                for column, value in enumerate(values, 1):
                    ws.cell(detail_row, column, value)
                ws.cell(detail_row, 4).number_format = _NUMBER_FORMAT
                ws.cell(detail_row, 6).number_format = _NUMBER_FORMAT
                detail_row += 1
        if not item.components:
            fallback = (
                "", "无组成明细", "计算结果", 0.0, "加",
                _display_amount(item.amount_minor, config.manifest.display_unit), "计算结果",
                item.verification_record_id or "未记录核验编号", "无", "无", "无", "无", "无", "无", "未展开",
            )
            for column, value in enumerate(fallback, 1):
                ws.cell(detail_row, column, value)
            detail_row += 1
        first_detail_row = header_row + 1
        calculated_row = detail_row
        statement_value_row = detail_row + 1
        difference_row = detail_row + 2
        conclusion_row = detail_row + 3
        ws.cell(calculated_row, 2, "规则计算结果")
        ws.cell(calculated_row, 6, f"=SUM(F{first_detail_row}:F{detail_row - 1})")
        ws.cell(statement_value_row, 2, "正表列示金额")
        ws.cell(statement_value_row, 6, f"='现金流量表'!C{statement_row}")
        ws.cell(difference_row, 2, "验证差异")
        ws.cell(difference_row, 6, f"=F{calculated_row}-F{statement_value_row}")
        ws.cell(conclusion_row, 2, "核对结论")
        ws.cell(conclusion_row, 6, f'=IF(ABS(F{difference_row})<0.005,"相符","不符")')
        for row_number in (calculated_row, statement_value_row, difference_row, conclusion_row):
            ws.cell(row_number, 2).font = Font(bold=True)
            if row_number != conclusion_row:
                ws.cell(row_number, 6).number_format = _NUMBER_FORMAT
        next_row = conclusion_row + 2

    statement = wb["现金流量表"]
    for item_id, statement_row in statement_rows.items():
        _set_internal_link(statement.cell(statement_row, 1), f"'计算验证底稿'!A{workpaper_rows[item_id]}")
    widths = {"A": 18, "B": 38, "C": 20, "D": 16, "E": 12, "F": 16, "G": 38, "H": 28, "I": 28, "J": 28, "K": 28, "L": 32, "M": 28, "N": 28, "O": 12}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A7"
    ws.sheet_view.showGridLines = False


def write_cashflow_statement(
    calculation: CalculationResult,
    prior_cashflow,
    config: RunConfig,
    validation: ValidationReport,
    statement_kind: str,
    output_path: Path,
    *,
    bridge=None,
    ledger_reconciliation=None,
    review_context: dict[str, object] | None = None,
) -> Path:
    if statement_kind == "最终" and not validation.is_clean:
        raise ValueError("完整性检查未通过，不得标记为最终正表")
    if statement_kind == "最终" and any(
        row.get("human_review_required")
        and row.get("confirmation_status") != "已确认"
        for row in (review_context or {}).get("decision_cases", ())
    ):
        raise ValueError("仍有未经人工确认的重大判断，不得标记为最终正表")
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active; ws.title = "现金流量表"
    ws.merge_cells("A1:D1"); ws["A1"] = "现金流量表"; ws["A1"].font = Font(name="宋体", size=16, bold=True); ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:D2"); ws["A2"] = f"{config.entity_name}　{config.period}　{statement_kind}正表"; ws["A2"].alignment = Alignment(horizontal="center")
    ws["A4"] = f"币种：{config.manifest.currency}"; ws["C4"] = f"单位：{config.manifest.display_unit}"
    for column, value in enumerate(("项目", "行次", "本期金额", "上期金额"), 1):
        cell = ws.cell(5, column, value); cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor="D9EAF7"); cell.alignment = Alignment(horizontal="center")
    priors = _prior_map(prior_cashflow, config.manifest.display_unit)
    prior_unmatched = []
    thin = Side(style="thin", color="B7B7B7")
    statement_rows = {}
    for row_number, item in enumerate(sorted(calculation.items, key=lambda x: x.display_order), 6):
        statement_rows[item.item_id] = row_number
        ws.cell(row_number, 1, item.name)
        ws.cell(row_number, 2, None)
        ws.cell(row_number, 3, float(from_minor_units(item.amount_minor, config.manifest.display_unit)))
        matched_prior = item.item_id in priors or item.name in priors
        prior = priors.get(item.item_id, priors.get(item.name, 0.0))
        if not matched_prior:
            prior_unmatched.append(item.name)
        ws.cell(row_number, 4, float(prior))
        for cell in ws[row_number]:
            cell.border = Border(bottom=thin)
        ws.cell(row_number, 3).number_format = _NUMBER_FORMAT
        ws.cell(row_number, 4).number_format = _NUMBER_FORMAT
    ws.column_dimensions["A"].width = 48; ws.column_dimensions["B"].width = 15; ws.column_dimensions["C"].width = 18; ws.column_dimensions["D"].width = 18
    ws.freeze_panes = "A6"; ws.sheet_view.showGridLines = False
    _write_verification_workpaper(
        wb,
        calculation,
        config,
        validation,
        statement_kind,
        statement_rows,
        bridge,
        ledger_reconciliation,
        tuple(prior_unmatched),
        review_context or {},
    )
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(output_path)
    check = load_workbook(output_path, data_only=False)
    if check.sheetnames != ["现金流量表", "计算验证底稿"] or check._external_links:
        check.close(); raise ValueError("输出文件结构不符合正表加可见验证底稿、无外链要求")
    if any(sheet.sheet_state != "visible" for sheet in check.worksheets):
        check.close(); raise ValueError("输出文件不得包含隐藏工作表")
    if any(isinstance(cell.value, str) and cell.value.startswith("=") for row in check["现金流量表"].iter_rows() for cell in row):
        check.close(); raise ValueError("输出正表不得包含计算公式")
    if any(
        isinstance(cell.value, str) and cell.value.startswith("=") and "[" in cell.value
        for sheet in check.worksheets for row in sheet.iter_rows() for cell in row
    ):
        check.close(); raise ValueError("输出文件不得包含外部工作簿公式")
    check.close()
    return output_path
